from datetime import datetime
import json 
import os 
from openpyxl import Workbook, load_workbook 
print("========================= Welcome to Ali's Daily Log System =========================")
folder = r"D:\01_Coding\03_Practice\06 Dailty Log"


def daily_life_tracker():
    while True:
        choice = input("Do You Want to write or Exit: ")
        if choice.lower().strip() == "exit":
            break
        timestamp = datetime.now().strftime(r"%Y:%m:%d %H:%M:%S")
        mood = input("How's Your Mode Today Buddy? ")
        win = input("What's Your Win Today? ")
        waste = input("What's Your Waste Today: ")
    
        # ========================= Save to JSON =========================
        json_file = r"daily_log.json"
        json_path = os.path.join(folder, json_file)
        entry =  {
            "Timestamp": timestamp,
            "Mood": mood,
            "Win": win,
            "Waste": waste 
        }
        with open(json_path, "a") as file:
            file.write(json.dumps(entry) + "\n")
        print("✅ Log Saved to JSON Successfully. ")
        
        # ========================= Save to Excel =========================
        excel_file = r"daily_log.xlsx"
        excel_path = os.path.join(folder, excel_file)
        if os.path.exists(excel_path):
            wb = load_workbook(excel_path)
            ws = wb.active 
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["TimeStamp", "Mood", "Win", "Waste"])
        ws.append([timestamp, mood, win, waste])
        wb.save(excel_path) 
        print("✅ Log Saved to Excel File Successfully. ")

        # ========================= Save to Text File =========================
        txt_file = r"daily_log.txt"
        txt_path = os.path.join(folder, txt_file)
        with open(txt_path, "a") as file:
            file.write(f"[{timestamp}] \nMood: {mood}\nWin: {win}\nWaste: {waste}")
            print("✅ Log Saved to Text File Successfully. ")


daily_life_tracker()
